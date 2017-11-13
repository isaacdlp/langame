# Copyright (c) 2017 Isaac de la Pena (isaacdlp@alum.mit.edu)
#
# Open-sourced according to the MIT LICENSE
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.


from random import choice
import warnings, sys, os, json, re, unicodedata
warnings.simplefilter("ignore")


file_base = "langame"
img_index = "EN"
lang_index = "ES"
lang_focus = None
word_focus = 0

action = "crop_pics"
if len(sys.argv) > 1:
    action = sys.argv[1]
    if len(sys.argv) > 2:
        lang_index = sys.argv[2].upper()
        if len(sys.argv) > 3:
            lang_focus = sys.argv[3].upper().split(",")
            map(lambda x: x.upper(), lang_focus)
            if "ALL" in lang_focus:
                lang_focus = None
            if lang_index in lang_focus:
                raise BaseException("The index language cannot be among the focus languages")
            if len(sys.argv) > 4:
                try:
                    word_focus = int(sys.argv[4])
                except:
                    print("Wrong word focus number")


genders = {
    4: "M",
    5: "F",
    6: "N"
}

inputs = {
    "M": "M, М",
    "F": "F, Ф",
    "N": "N, Н"
}

note_reg = re.compile("\([^\)]+\)")
clean_reg = re.compile("[^\w]+")
normal_reg = re.compile("[À-ÿьъ]+")
repeat_reg = re.compile("((\w)\\2{1,})")
replacements = (("й", "и"), ("ы", "и"), ("щ", "ш"), ("в", "б"))

def do_clean(text):
    text = note_reg.sub("", text).lower()
    text = clean_reg.sub("", text)
    return text

def do_normal(text):
    for raw in normal_reg.findall(text):
        normal = unicodedata.normalize("NFD", raw).encode("ascii", "ignore").decode("utf-8")
        text = text.replace(raw, normal)
    for replacement in replacements:
        text = text.replace(replacement[0], replacement[1])
    for repeat in repeat_reg.findall(text):
        text = text.replace(repeat[0], repeat[1])
    return text

def do_eval(guess, solutions):
    match = "WRONG"
    points = 0
    clean_guess = do_clean(guess)
    normal_guess = do_normal(clean_guess)
    for solution in solutions.split(", "):
        clean_sol = do_clean(solution)
        if clean_guess == clean_sol:
            match = "RIGHT"
            points = 0.5
            break
        else:
            normal_sol = do_normal(clean_sol)
            if normal_guess == normal_sol:
                match = "ALMOST"
                points = 0.25
    return match, points


if action == "update":

    from openpyxl import load_workbook

    # Update the JSON dictionary

    wb = load_workbook(filename="%s.xlsx" % file_base)
    sh = wb[wb.sheetnames[0]]
    langs = []

    col = 0
    while True:
        col += 1
        lang = sh.cell(row=1, column=col).value
        if lang is None:
            break
        else:
            langs.append(lang.upper())

    unique = dict([(lang, []) for lang in langs])
    words = []
    concept = None

    row = 1
    while True:
        row += 1
        if concept is not None:
            if len(word) > 1:
                for lang, sym in word.items():
                    if sym in unique[lang]:
                        print("Warning duplicate in %s: %s" % (lang, word[lang]))
                    unique[lang].append(sym)
                words.append(word)
        concept = sh.cell(row=row, column=1).value
        if concept is None:
            break
        word = {}
        for col, lang in enumerate(langs):
            con = []
            col += 1
            cell = sh.cell(row=row, column=col)
            sym = cell.value
            if sym is not None:
                con.append(sym)
                gender = cell.fill.start_color.index
                gen = genders[gender] if gender in genders else None
                if gen is not None:
                    con.append(gen)
                word[lang] = con

    print("Writing %i words to JSON" % len(words))
    with open("%s.json" % file_base, "w") as f:
        json.dump(words, f)

elif action == "get_pics":

    from selenium import webdriver as web
    from selenium.webdriver.chrome.options import Options
    import requests as req

    with open("%s.json" % file_base, "r") as f:
        words = json.load(f)

    search_url = "https://www.flickr.com/search/?license=4,5,9,10&dimension_search_mode=min&height=1024&width=1024&media=photos&text=%s"
    view_url = "%ssizes/l"

    options = Options()
    options.add_argument("--disable-notifications")
    browser = web.Chrome(chrome_options=options)
    browser.set_window_size(1000, 1000)

    for word in words:
        if img_index in word:
            sym = word[img_index][0]
            sym = sym.split(", ")[0]
            filename = "%s/%s" % (file_base, sym)
            if not os.path.isfile("%s.jpg" % filename):
                browser.get(search_url % do_clean(sym))
                links = browser.find_elements_by_css_selector("div.photo-list-photo-interaction a.overlay")
                if len(links) > 0:
                    if len(links) > 10:
                        links = links[:10]
                    tags = []
                    for link in links:
                        tags.append(link.get_attribute("href"))
                    for num, tag in enumerate(tags):
                        browser.get(view_url % tag)
                        img = browser.find_element_by_css_selector("div#allsizes-photo img")
                        src = img.get_attribute("src")
                        res = req.get(src, stream=True)
                        imgname = "%s" % filename
                        if num > 0:
                            imgname = "%s-%i" % (imgname, num)
                        with open("%s.jpg" % imgname, 'wb') as f:
                            f.write(res.raw.data)
                else:
                    print("No pictures for '%s'" % sym)

    browser.quit()

elif action == "crop_pics":

    # Crop pictures

    from PIL import Image

    img_size = 1024

    for file in os.listdir(file_base):
        if file.endswith(".jpg"):
            img_file = "%s/%s" % (file_base, file)
            im = Image.open(img_file)
            width, height = im.size  # Get dimensions
            if width != height:
                img_width = width if width < img_size else img_size
                img_height = height if height < img_size else img_size
                img_width = min((img_width, img_height))
                img_height = img_width
                left = (width - img_width) / 2
                top = (height - img_height) / 2
                right = (width + img_width) / 2
                bottom = (height + img_height) / 2
                crop = im.crop((left, top, right, bottom))
                crop.save(img_file, "JPEG")
                print("Cropped %s" % file)

else:

    # Play the game!

    with open("%s.json" % file_base, "r") as f:
        words = json.load(f)

    if word_focus > 0 and not word_focus > len(words):
        words = words[-word_focus:]
    used = []

    rounds = 0
    score = 0.0
    while True:
        rounds += 1
        concept = None
        lang = None
        sym = None
        gen = None
        for i in range(100):
            word = choice(words)
            # In Python 3 keys() is an iterator, not a list
            langs = list(word.keys())
            if lang_index in langs:
                concept = word[lang_index][0]
                langs.remove(lang_index)
                lang = choice(langs).upper()
                if lang_focus is None or lang in lang_focus:
                    pair = "%s@%s" % (concept, lang)
                    if pair not in used:
                        used.append(pair)
                        con = word[lang]
                        sym = con[0]
                        gen = con[1] if len(con) > 1 else None
                        break

        if sym is None:
            print("Not enough words!")
            break

        print("Round %i [%.2f] '%s' in %s" % (rounds, score, concept, lang))
        guess = input()
        if guess == "@":
            break
        result, points = do_eval(guess, sym)
        print("%s! It is '%s'" % (result, sym))
        if gen is not None:
            print("Which gender is it? [M]asculine, [N]eutral or [F]emenine?")
            guess = input()
            result, pts = do_eval(guess, inputs[gen])
            print("%s! It is '%s'" % (result, gen))
            points += pts
        else:
            points *= 2
        score += points

    rounds -= 1
    accuracy = (score / rounds) * 100 if rounds > 0 else 0
    print("Total points [%.2f] rounds [%i] accuracy [%.2f%%]" % (score, rounds, accuracy))
    print("Thank you for playing!")